import os
from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from openpyxl import load_workbook


class File:

    def __init__(self, path=''):
        self.path = path
        self.pgn_list = []
        self.pgn_set = set()
        self.str_list = []

    def __str__(self):
        return 'Cl2000 Log File'


def create_dict():
    wb = load_workbook(filename='j1939_pgn_list.xlsx')
    ws = wb.active

    dictionary = {}
    for row in range(3, ws.max_row + 1):
        key = ws.cell(row, 2).value
        value = ws.cell(row, 3).value
        dictionary[key] = value

    return dictionary


def get_path():
    data_log.path = filedialog.askopenfilename(filetypes=(("txt", "*.txt"), ("all files", "*.*")))
    file_path_label.config(text=data_log.path[-65:])
    get_content()


def get_content():
    # Read Content From File
    try:
        f = open(data_log.path, 'r')
        all_content = f.readlines()
        f.close()
    except PermissionError:
        messagebox.showerror('Error', 'No file selected!')
        raise Exception('No file selected!')

    pgn_list = []
    for line in all_content:
        if line[0] == '2':
            pgn_list.append(line.rstrip())

    pgn_set = set()
    for num in pgn_list:
        pgn = num[15:23]
        if pgn[-1] == ';':
            pgn_set.add(pgn[1:5])
        else:
            pgn_set.add(pgn[2:6])

    hex_set = []
    dec_set = []
    for num in pgn_set:
        x = int(num, 16)
        dec_set.append(x)
        hex_set.append(num)

    hex_set.sort()
    dec_set.sort()

    str_list = ''
    for index in dec_set:
        if index != 0:
            if index in pgn_dict:
                str_list += '{' + str(index) + ',kvFalse}/* ' + pgn_dict[index][:33] + ' */\n'
            else:
                str_list += '{' + str(index) + ',kvFalse}/* Proprietary */\n'

    output_text.delete("1.0", "end")
    data_log.str_list = str_list
    output_text.insert('end', str_list)


def save_to_file():
    file = filedialog.asksaveasfilename(defaultextension='.txt',
                                        filetypes=(("txt", "*.txt"), ("all files", "*.*")))
    with open(file, "w") as f:
        for index in data_log.str_list:
            f.write(index)


# Set data_log as File class
data_log = File()

# Tkinter Window Setup
window = Tk()

# Create Variables
current_dir = os.getcwd()
pgn_dict = create_dict()

# Set window properties
window.title('CL2000 Log Unloader')
# window.geometry("455x675")
window.iconbitmap(current_dir + '/images/favicon.png')


# Configure Grid
window.columnconfigure(0, weight=1)

# Create Labels for browsing Files and Displaying File Path
label_file_explorer = Label(window, text="Decode CL2000 Log Files", width=30, height=2)
label_file_explorer.grid(column=0, row=0)

version_label = Label(window, text="Version 1.1.0", width=30, height=2)
version_label.grid(column=0, row=5)

file_path_label = Label(window, text=data_log.path, background='white', width=52, borderwidth=1, relief='sunken')
file_path_label.grid(column=0, row=1, sticky=W, pady=2, padx=5, ipady=3)

# Output box for converter PGNs
output_text = Text(window, width=55, height=35)
output_text.grid(column=0, row=2, padx=5, columnspan=3)

# Setup Buttons
# Browse Button
button_explore = Button(window, text="Browse Files", command=get_path)
button_explore.grid(column=0, row=1, sticky=E, padx=5)

# Save Button
button_save_list = Button(window, text="Save", command=save_to_file)
button_save_list.grid(column=0, row=5, pady=5, padx=5, sticky=W, ipady=5, ipadx=30)

# Exit Button
button_exit = Button(window, text="Exit", command=window.destroy)
button_exit.grid(column=0, row=5, pady=5, padx=5, sticky=E, ipady=5, ipadx=30)

window.resizable(False, False)
# Let the window wait for any events
window.mainloop()
